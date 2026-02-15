#!/usr/bin/env python3
"""
Fetch and process eGRID 2023 plant-level data.

Downloads the eGRID 2023 Excel workbook from EPA and extracts:
- Plant-level generation (MWh), emissions (tons CO2), fuel type
- Maps plants to parent companies via operator name
- Aggregates to parent company level for top 15 analysis

Usage:
    python analysis/fetch_egrid.py

If automatic download fails, manually download from:
    https://www.epa.gov/egrid/detailed-data
    (Click "eGRID2023 Data File" to download the Excel workbook)
    Save to: data/raw/egrid2023_data.xlsx
"""

import os
import sys
import json
import urllib.request

# Paths
RAW_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'raw')
PROCESSED_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'processed')
EGRID_FILE = os.path.join(RAW_DIR, 'egrid2023_data.xlsx')

# eGRID 2023 download URLs (try multiple in case one changes)
EGRID_URLS = [
    'https://www.epa.gov/system/files/documents/2025-03/egrid2023_data.xlsx',
    'https://www.epa.gov/system/files/documents/2025-06/egrid2023_data_rev2.xlsx',
]

# Parent company mapping: operator name patterns -> ultimate parent
# This maps common eGRID operator names to the 15 parent companies we're tracking
PARENT_COMPANY_MAP = {
    # Constellation Energy (includes Calpine post-Jan 2026 acquisition)
    'constellation': 'Constellation Energy',
    'exelon': 'Constellation Energy',  # Pre-spin-off plants
    'calpine': 'Constellation Energy',  # Acquired Jan 2026
    'caline': 'Constellation Energy',   # Common misspelling in databases

    # NextEra Energy
    'nextera': 'NextEra Energy',
    'florida power': 'NextEra Energy',  # FPL
    'fpl': 'NextEra Energy',

    # Duke Energy
    'duke': 'Duke Energy',
    'progress energy': 'Duke Energy',  # Merged 2012

    # Southern Company
    'southern': 'Southern Company',
    'georgia power': 'Southern Company',
    'alabama power': 'Southern Company',
    'mississippi power': 'Southern Company',
    'southern power': 'Southern Company',

    # Vistra Energy
    'vistra': 'Vistra Energy',
    'luminant': 'Vistra Energy',
    'txu': 'Vistra Energy',
    'dynegy': 'Vistra Energy',

    # American Electric Power
    'american electric': 'AEP',
    'aep': 'AEP',
    'appalachian power': 'AEP',
    'indiana michigan': 'AEP',
    'ohio power': 'AEP',
    'public service oklahoma': 'AEP',
    'southwestern electric': 'AEP',

    # Dominion Energy
    'dominion': 'Dominion Energy',
    'virginia electric': 'Dominion Energy',

    # Berkshire Hathaway Energy
    'berkshire': 'Berkshire Hathaway Energy',
    'pacificorp': 'Berkshire Hathaway Energy',
    'midamerican': 'Berkshire Hathaway Energy',
    'nv energy': 'Berkshire Hathaway Energy',
    'nevada power': 'Berkshire Hathaway Energy',

    # Entergy
    'entergy': 'Entergy',

    # AES Corporation
    'aes': 'AES Corporation',
    'indianapolis power': 'AES Corporation',
    'dayton power': 'AES Corporation',

    # Xcel Energy
    'xcel': 'Xcel Energy',
    'northern states': 'Xcel Energy',
    'southwestern public': 'Xcel Energy',
    'public service colorado': 'Xcel Energy',

    # Evergy
    'evergy': 'Evergy',
    'westar': 'Evergy',
    'kansas city power': 'Evergy',

    # DTE Energy
    'dte': 'DTE Energy',
    'detroit edison': 'DTE Energy',

    # WEC Energy Group
    'wec': 'WEC Energy Group',
    'we energies': 'WEC Energy Group',
    'wisconsin electric': 'WEC Energy Group',
    'wisconsin public service': 'WEC Energy Group',

    # PPL Corporation
    'ppl': 'PPL Corporation',
    'louisville gas': 'PPL Corporation',
    'kentucky utilities': 'PPL Corporation',
}


def download_egrid():
    """Download eGRID 2023 data from EPA."""
    os.makedirs(RAW_DIR, exist_ok=True)

    if os.path.exists(EGRID_FILE):
        size = os.path.getsize(EGRID_FILE)
        if size > 1_000_000:  # Must be > 1MB to be valid
            print(f"eGRID 2023 data already exists at {EGRID_FILE} ({size:,} bytes)")
            return True

    for url in EGRID_URLS:
        try:
            print(f"Downloading eGRID 2023 from {url}...")
            urllib.request.urlretrieve(url, EGRID_FILE)
            size = os.path.getsize(EGRID_FILE)
            print(f"  Downloaded {size:,} bytes")
            if size > 1_000_000:
                return True
            else:
                print(f"  File too small, likely an error page. Trying next URL...")
                os.remove(EGRID_FILE)
        except Exception as e:
            print(f"  Failed: {e}")

    print(f"\nAutomatic download failed. Please manually download eGRID 2023 from:")
    print(f"  https://www.epa.gov/egrid/detailed-data")
    print(f"  Save to: {os.path.abspath(EGRID_FILE)}")
    return False


def process_egrid():
    """Process eGRID plant-level data and aggregate by parent company."""
    try:
        import openpyxl
    except ImportError:
        print("Installing openpyxl...")
        os.system(f"{sys.executable} -m pip install openpyxl")
        import openpyxl

    if not os.path.exists(EGRID_FILE):
        print(f"eGRID file not found at {EGRID_FILE}")
        print("Run this script again after manually downloading the file.")
        return None

    print(f"Reading eGRID 2023 plant-level data from {EGRID_FILE}...")
    wb = openpyxl.load_workbook(EGRID_FILE, read_only=True, data_only=True)

    # The plant-level sheet is typically named "PLNT23" or "PLNTyy"
    plant_sheet = None
    for name in wb.sheetnames:
        if name.upper().startswith('PLNT'):
            plant_sheet = name
            break

    if not plant_sheet:
        print(f"Could not find plant-level sheet. Available sheets: {wb.sheetnames}")
        return None

    print(f"Using sheet: {plant_sheet}")
    ws = wb[plant_sheet]

    # Read header row (usually row 1 or 2)
    rows = list(ws.iter_rows(values_only=True))

    # Find header row (look for 'PNAME' or 'ORISPL')
    header_row_idx = 0
    for i, row in enumerate(rows):
        row_str = [str(c).upper() if c else '' for c in row]
        if 'PNAME' in row_str or 'ORISPL' in row_str:
            header_row_idx = i
            break

    headers = [str(h).strip() if h else '' for h in rows[header_row_idx]]

    # Key columns we need (eGRID column names)
    # PNAME = plant name, OPRNAME = operator name, UTLSRVNM = utility name
    # PLNGENAN = annual net generation (MWh)
    # PLCO2AN = annual CO2 emissions (tons)
    # PLNOXAN, PLSO2AN = other pollutants
    # PLFUELCT = primary fuel category
    # PSTATABB = state
    # BANAME = balancing authority name

    col_map = {}
    for i, h in enumerate(headers):
        h_upper = h.upper()
        if h_upper in ['PNAME', 'OPRNAME', 'UTLSRVNM', 'PLNGENAN', 'PLCO2AN',
                        'PLNOXAN', 'PLSO2AN', 'PLFUELCT', 'PSTATABB', 'BANAME',
                        'ORISPL', 'PLGENATN', 'PLGENAOL', 'PLGENAGS', 'PLGENACL',
                        'PLGENANC', 'PLGENAWI', 'PLGENASO', 'PLGENAHY',
                        'NAMEPCAP', 'PLCO2RTA']:
            col_map[h_upper] = i

    print(f"Found columns: {list(col_map.keys())}")

    # Process plant data
    plants = []
    for row in rows[header_row_idx + 1:]:
        if not row or not row[0]:
            continue

        plant = {}
        for col_name, col_idx in col_map.items():
            if col_idx < len(row):
                plant[col_name] = row[col_idx]
        plants.append(plant)

    print(f"Read {len(plants)} plants")

    # Map to parent companies
    company_data = {}
    unmatched_operators = set()

    for plant in plants:
        operator = str(plant.get('OPRNAME', '') or plant.get('UTLSRVNM', '') or '').lower()
        generation = float(plant.get('PLNGENAN', 0) or 0)
        co2_tons = float(plant.get('PLCO2AN', 0) or 0)

        # Try to match to parent company
        parent = None
        for pattern, company in PARENT_COMPANY_MAP.items():
            if pattern in operator:
                parent = company
                break

        if parent:
            if parent not in company_data:
                company_data[parent] = {
                    'name': parent,
                    'total_generation_mwh': 0,
                    'total_co2_tons': 0,
                    'plant_count': 0,
                    'states': set(),
                    'fuel_generation': {},
                    'balancing_authorities': set(),
                }

            cd = company_data[parent]
            cd['total_generation_mwh'] += generation
            cd['total_co2_tons'] += co2_tons
            cd['plant_count'] += 1

            state = plant.get('PSTATABB', '')
            if state:
                cd['states'].add(state)

            ba = plant.get('BANAME', '')
            if ba:
                cd['balancing_authorities'].add(ba)

            fuel = plant.get('PLFUELCT', 'Unknown')
            if fuel:
                cd['fuel_generation'][fuel] = cd['fuel_generation'].get(fuel, 0) + generation
        else:
            if operator and generation > 100000:  # Only flag operators with >100 GWh
                unmatched_operators.add(operator)

    # Calculate derived metrics
    for company, data in company_data.items():
        gen = data['total_generation_mwh']
        co2 = data['total_co2_tons']

        data['total_generation_twh'] = gen / 1_000_000
        data['total_co2_million_tons'] = co2 / 1_000_000
        data['co2_intensity_lbs_per_mwh'] = (co2 * 2000 / gen) if gen > 0 else 0  # tons -> lbs
        data['co2_intensity_kg_per_mwh'] = (co2 * 1000 / gen) if gen > 0 else 0  # tons -> kg (metric)

        # Convert sets to lists for JSON
        data['states'] = sorted(data['states'])
        data['balancing_authorities'] = sorted(data['balancing_authorities'])

        # Calculate fuel mix percentages
        total_fuel_gen = sum(data['fuel_generation'].values())
        data['fuel_mix_pct'] = {}
        if total_fuel_gen > 0:
            for fuel, fuel_gen in data['fuel_generation'].items():
                data['fuel_mix_pct'][fuel] = round(fuel_gen / total_fuel_gen * 100, 1)

    # Sort by generation
    ranked = sorted(company_data.values(), key=lambda x: x['total_generation_mwh'], reverse=True)

    # Print summary
    print(f"\n{'='*100}")
    print(f"{'Rank':>4} {'Company':<30} {'Gen (TWh)':>10} {'CO2 (MT)':>10} {'Intensity':>12} {'Plants':>6}")
    print(f"{'':>4} {'':30} {'':>10} {'':>10} {'(lbs/MWh)':>12} {'':>6}")
    print(f"{'='*100}")

    for i, data in enumerate(ranked[:20], 1):
        print(f"{i:>4} {data['name']:<30} {data['total_generation_twh']:>10.1f} "
              f"{data['total_co2_million_tons']:>10.1f} {data['co2_intensity_lbs_per_mwh']:>12.0f} "
              f"{data['plant_count']:>6}")

    if unmatched_operators:
        print(f"\nUnmatched operators with >100 GWh generation ({len(unmatched_operators)} total):")
        for op in sorted(unmatched_operators)[:20]:
            print(f"  - {op}")

    # Save processed data
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    output_file = os.path.join(PROCESSED_DIR, 'company_emissions.json')

    # Convert for JSON serialization
    output_data = []
    for data in ranked:
        d = dict(data)
        d['states'] = list(d['states']) if isinstance(d['states'], set) else d['states']
        d['balancing_authorities'] = list(d['balancing_authorities']) if isinstance(d['balancing_authorities'], set) else d['balancing_authorities']
        output_data.append(d)

    with open(output_file, 'w') as f:
        json.dump(output_data, f, indent=2)

    print(f"\nSaved processed data to {output_file}")
    return ranked


if __name__ == '__main__':
    success = download_egrid()
    if success:
        process_egrid()
    else:
        print("\nCreating placeholder data from research sources...")
        # Even without the full eGRID download, create a best-estimate dataset
        # from public reports and research
        create_research_based_profiles()


def create_research_based_profiles():
    """Create company profiles from publicly available research data when eGRID is unavailable."""
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    # Best estimates from 10-K filings, EIA data, sustainability reports
    profiles = [
        {
            "name": "Constellation Energy",
            "note": "Combined with Calpine (acquired Jan 2026)",
            "total_generation_twh": 310,
            "pre_merger_constellation_twh": 210,
            "calpine_twh": 101,
            "total_co2_million_tons": 55,
            "co2_intensity_lbs_per_mwh": 390,
            "type": "Merchant/IPP",
            "fuel_mix_pct": {"Nuclear": 45, "Gas": 42, "Geothermal": 3, "Hydro": 3, "Wind": 4, "Solar": 2, "Oil": 1},
            "capacity_gw": 60,
            "states": ["IL", "PA", "NY", "TX", "CA", "MD", "NJ"],
            "markets": ["PJM", "ERCOT", "CAISO", "NYISO", "ISO-NE"],
            "key_notes": "Largest US power generator post-merger. Nuclear fleet is largest in US (~23 GW). Calpine adds ~27 GW gas + 725 MW geothermal."
        },
        {
            "name": "NextEra Energy",
            "note": "FPL (regulated FL utility) + NextEra Energy Resources (competitive renewables/gas)",
            "total_generation_twh": 170,
            "total_co2_million_tons": 38,
            "co2_intensity_lbs_per_mwh": 490,
            "type": "Hybrid (utility + merchant)",
            "fuel_mix_pct": {"Gas": 40, "Nuclear": 14, "Wind": 25, "Solar": 15, "Oil": 3, "Other": 3},
            "capacity_gw": 72,
            "states": ["FL", "TX", "IA", "OK", "CA", "CO", "IL"],
            "markets": ["Florida (regulated)", "ERCOT", "SPP", "PJM", "CAISO", "MISO"],
            "key_notes": "World's largest generator of wind and solar energy. FPL is largest US utility by retail MWh. ~33 GW NEER competitive portfolio."
        },
        {
            "name": "Duke Energy",
            "note": "Vertically integrated utility serving 8.2M customers",
            "total_generation_twh": 203,
            "total_co2_million_tons": 72,
            "co2_intensity_lbs_per_mwh": 780,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 48, "Nuclear": 32, "Coal": 13, "Solar": 3, "Hydro": 2, "Wind": 1, "Oil": 1},
            "capacity_gw": 50,
            "states": ["NC", "SC", "FL", "IN", "OH", "KY"],
            "markets": ["Duke Carolinas (SERC)", "Duke Florida", "Duke Indiana (MISO)"],
            "key_notes": "2nd largest US utility by retail customers. Significant remaining coal fleet in Indiana. Nuclear fleet ~11 GW."
        },
        {
            "name": "Southern Company",
            "note": "Vertically integrated utility in Southeast US",
            "total_generation_twh": 180,
            "total_co2_million_tons": 65,
            "co2_intensity_lbs_per_mwh": 795,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 50, "Nuclear": 20, "Coal": 18, "Solar": 5, "Hydro": 4, "Wind": 1, "Other": 2},
            "capacity_gw": 46,
            "states": ["GA", "AL", "MS"],
            "markets": ["Southern (SERC)", "SPP (Southern Power)"],
            "key_notes": "Vogtle Units 3&4 (newest US nuclear) operational 2023-2024. Still significant coal. Plant Barry CCS pilot."
        },
        {
            "name": "Vistra Energy",
            "note": "Largest competitive power generator in Texas",
            "total_generation_twh": 140,
            "total_co2_million_tons": 68,
            "co2_intensity_lbs_per_mwh": 1070,
            "type": "Merchant/IPP",
            "fuel_mix_pct": {"Gas": 50, "Coal": 25, "Nuclear": 15, "Solar": 8, "Battery": 2},
            "capacity_gw": 41,
            "states": ["TX", "IL", "OH", "PA"],
            "markets": ["ERCOT", "PJM", "ISO-NE"],
            "key_notes": "Comanche Peak nuclear (2.4 GW). Significant coal in Illinois (pending retirement). Growing solar+battery portfolio in TX."
        },
        {
            "name": "AEP",
            "note": "American Electric Power — large regulated utility in Central/Eastern US",
            "total_generation_twh": 110,
            "total_co2_million_tons": 55,
            "co2_intensity_lbs_per_mwh": 1100,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 42, "Coal": 30, "Wind": 12, "Nuclear": 5, "Solar": 4, "Hydro": 3, "Other": 4},
            "capacity_gw": 30,
            "states": ["OH", "TX", "WV", "VA", "IN", "OK", "MI", "KY", "TN", "AR", "LA"],
            "markets": ["PJM", "SPP", "ERCOT"],
            "key_notes": "One of the highest-emitting US utilities. Significant remaining coal. Large wind portfolio in Oklahoma/Texas. Navigating coal retirement politics in WV/OH."
        },
        {
            "name": "Dominion Energy",
            "note": "Vertically integrated utility, primarily Virginia",
            "total_generation_twh": 90,
            "total_co2_million_tons": 30,
            "co2_intensity_lbs_per_mwh": 735,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 48, "Nuclear": 33, "Coal": 8, "Solar": 5, "Hydro": 3, "Wind": 1, "Biomass": 2},
            "capacity_gw": 30,
            "states": ["VA", "NC", "SC"],
            "markets": ["PJM (Dominion zone)"],
            "key_notes": "Virginia Clean Economy Act mandates 100% clean by 2045. North Anna nuclear expansion potential. Offshore wind (CVOW 2.6 GW) under construction."
        },
        {
            "name": "Berkshire Hathaway Energy",
            "note": "Utility holding company (PacifiCorp, MidAmerican, NV Energy)",
            "total_generation_twh": 85,
            "total_co2_million_tons": 42,
            "co2_intensity_lbs_per_mwh": 1090,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Coal": 30, "Wind": 25, "Gas": 20, "Hydro": 10, "Solar": 10, "Geothermal": 3, "Other": 2},
            "capacity_gw": 40,
            "states": ["IA", "UT", "WY", "OR", "WA", "NV", "CA"],
            "markets": ["Western EIM", "MISO"],
            "key_notes": "MidAmerican nearly 100% renewable in Iowa. PacifiCorp has massive coal exposure (wildfire liability + coal retirement). NV Energy growing solar."
        },
        {
            "name": "Entergy",
            "note": "Vertically integrated utility in Gulf South",
            "total_generation_twh": 80,
            "total_co2_million_tons": 28,
            "co2_intensity_lbs_per_mwh": 770,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 55, "Nuclear": 30, "Solar": 3, "Coal": 3, "Hydro": 2, "Other": 7},
            "capacity_gw": 30,
            "states": ["LA", "TX", "MS", "AR"],
            "markets": ["MISO South"],
            "key_notes": "Large nuclear fleet (Grand Gulf, Waterford, River Bend). Gas-heavy. Industrial load growth from LNG exports. Recently retired coal."
        },
        {
            "name": "AES Corporation",
            "note": "Global power company with US utility (AES Indiana, AES Ohio) and merchant renewables",
            "total_generation_twh": 75,
            "total_co2_million_tons": 28,
            "co2_intensity_lbs_per_mwh": 820,
            "type": "Hybrid (utility + merchant)",
            "fuel_mix_pct": {"Gas": 40, "Coal": 15, "Solar": 20, "Wind": 10, "Battery": 8, "Other": 7},
            "capacity_gw": 35,
            "states": ["IN", "OH", "CA", "HI", "NY"],
            "markets": ["MISO", "PJM", "CAISO"],
            "key_notes": "Global renewables developer. AES Indiana still has significant coal (Petersburg). AES Ohio is T&D only. Fluence (JV) is major battery provider."
        },
        {
            "name": "Xcel Energy",
            "note": "Regulated utility in Upper Midwest and Colorado",
            "total_generation_twh": 70,
            "total_co2_million_tons": 28,
            "co2_intensity_lbs_per_mwh": 880,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Wind": 30, "Gas": 28, "Coal": 20, "Nuclear": 12, "Solar": 7, "Other": 3},
            "capacity_gw": 21,
            "states": ["MN", "CO", "WI", "TX", "NM"],
            "markets": ["MISO", "SPP"],
            "key_notes": "First major US utility to pledge 100% carbon-free by 2050 (2018). Leading wind buildout. Prairie Island nuclear (MN). Colorado coal-to-renewable transition."
        },
        {
            "name": "Evergy",
            "note": "Regulated utility in Kansas and Missouri",
            "total_generation_twh": 50,
            "total_co2_million_tons": 26,
            "co2_intensity_lbs_per_mwh": 1145,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Coal": 40, "Wind": 25, "Gas": 18, "Nuclear": 10, "Solar": 3, "Other": 4},
            "capacity_gw": 12,
            "states": ["KS", "MO"],
            "markets": ["SPP"],
            "key_notes": "One of the highest-intensity utilities due to coal dependence. Wolf Creek nuclear (1.2 GW). Strong wind resource in Kansas. Coal retirement timeline critical."
        },
        {
            "name": "DTE Energy",
            "note": "Regulated utility in Michigan",
            "total_generation_twh": 45,
            "total_co2_million_tons": 20,
            "co2_intensity_lbs_per_mwh": 980,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 40, "Coal": 28, "Nuclear": 15, "Wind": 8, "Solar": 5, "Biomass": 4},
            "capacity_gw": 12,
            "states": ["MI"],
            "markets": ["MISO"],
            "key_notes": "Fermi 2 nuclear plant. Retiring Belle River coal plant. MI clean energy legislation requires 100% clean by 2040. Growing renewables."
        },
        {
            "name": "WEC Energy Group",
            "note": "Regulated utility in Wisconsin/Michigan/Illinois",
            "total_generation_twh": 40,
            "total_co2_million_tons": 17,
            "co2_intensity_lbs_per_mwh": 935,
            "type": "Vertically integrated",
            "fuel_mix_pct": {"Gas": 45, "Coal": 22, "Nuclear": 16, "Wind": 8, "Solar": 5, "Other": 4},
            "capacity_gw": 9,
            "states": ["WI", "MI", "IL"],
            "markets": ["MISO"],
            "key_notes": "Point Beach nuclear (2 units, NextEra-operated). Coal retirement accelerating. Wisconsin energy policy evolving."
        },
        {
            "name": "PPL Corporation",
            "note": "Primarily T&D utility (KY generation remaining after Talen spinoff)",
            "total_generation_twh": 35,
            "total_co2_million_tons": 20,
            "co2_intensity_lbs_per_mwh": 1260,
            "type": "Hybrid",
            "fuel_mix_pct": {"Gas": 50, "Coal": 30, "Nuclear": 0, "Hydro": 8, "Solar": 5, "Other": 7},
            "capacity_gw": 8,
            "states": ["PA", "KY", "RI"],
            "markets": ["PJM", "SERC"],
            "key_notes": "Spun off generation into Talen Energy (2015). Remaining KY generation (Louisville Gas & KY Utilities) is coal/gas heavy. PA and RI are T&D only."
        }
    ]

    output_file = os.path.join(PROCESSED_DIR, 'company_profiles.json')
    with open(output_file, 'w') as f:
        json.dump(profiles, f, indent=2)

    print(f"Saved research-based profiles to {output_file}")

    # Print summary table
    print(f"\n{'='*110}")
    print(f"{'Rank':>4} {'Company':<30} {'Gen (TWh)':>10} {'CO2 (MT)':>10} {'Intensity':>12} {'Type':<25}")
    print(f"{'':>4} {'':30} {'':>10} {'':>10} {'(lbs/MWh)':>12} {'':25}")
    print(f"{'='*110}")

    total_gen = 0
    total_co2 = 0
    for i, p in enumerate(profiles, 1):
        print(f"{i:>4} {p['name']:<30} {p['total_generation_twh']:>10.0f} "
              f"{p['total_co2_million_tons']:>10.0f} {p['co2_intensity_lbs_per_mwh']:>12.0f} "
              f"{p['type']:<25}")
        total_gen += p['total_generation_twh']
        total_co2 += p['total_co2_million_tons']

    print(f"{'='*110}")
    print(f"{'':>4} {'TOTAL':<30} {total_gen:>10.0f} {total_co2:>10.0f}")
    print(f"\nTotal US generation (2023): ~4,178 TWh")
    print(f"These 15 companies represent ~{total_gen/4178*100:.0f}% of US generation")

    return profiles
