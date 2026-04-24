#!/usr/bin/env python3
"""
Extract 2023 baseline CO2 emissions by ISO from eGRID 2023 plant-level data.

Downloads or reads the eGRID 2023 Excel workbook, sums PLCO2AN (annual CO2
emissions in short tons) by balancing authority, converts to metric tons,
and outputs per-ISO baseline emissions for the SMARTargets AT trajectory.

Also extracts per-fuel emission rates to update egrid_emission_rates.json
with precise plant-level aggregations (replacing rounded estimates for
MISO/SPP identified in the CO2 methodology audit).

Usage:
    python scripts/step0_extract_egrid_baselines.py
    python scripts/step0_extract_egrid_baselines.py --update-rates

Prerequisites:
    pip install openpyxl

If auto-download fails, manually download from:
    https://www.epa.gov/egrid/detailed-data
    Save to: data/raw/egrid2023_data.xlsx
"""

import os
import sys
import json
import urllib.request

# Paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.join(SCRIPT_DIR, '..')
RAW_DIR = os.path.join(ROOT_DIR, 'data', 'raw')
OUTPUT_FILE = os.path.join(ROOT_DIR, 'data', 'egrid_2023_baseline_emissions.json')
EGRID_FILE = os.path.join(RAW_DIR, 'egrid2023_data.xlsx')

# eGRID 2023 download URLs
EGRID_URLS = [
    'https://www.epa.gov/system/files/documents/2025-03/egrid2023_data.xlsx',
    'https://www.epa.gov/system/files/documents/2025-06/egrid2023_data_rev2.xlsx',
]

# BA code → ISO mapping
BA_TO_ISO = {
    'CISO': 'CAISO',
    'ERCO': 'ERCOT',
    'PJM':  'PJM',
    'NYIS': 'NYISO',
    'ISNE': 'NEISO',
    'MISO': 'MISO',
    'SWPP': 'SPP',
}

SHORT_TONS_TO_METRIC_TONS = 1 / 1.10231


def download_egrid():
    """Download eGRID 2023 Excel file if not present."""
    os.makedirs(RAW_DIR, exist_ok=True)
    if os.path.exists(EGRID_FILE):
        print(f"eGRID file already exists: {EGRID_FILE}")
        return True

    for url in EGRID_URLS:
        try:
            print(f"Downloading from {url}...")
            urllib.request.urlretrieve(url, EGRID_FILE)
            size_mb = os.path.getsize(EGRID_FILE) / 1e6
            print(f"Downloaded: {size_mb:.1f} MB")
            return True
        except Exception as e:
            print(f"Failed: {e}")

    print(f"\nAuto-download failed. Please manually download eGRID 2023 from:")
    print(f"  https://www.epa.gov/egrid/detailed-data")
    print(f"  Save to: {EGRID_FILE}")
    return False


def extract_plant_emissions():
    """Read plant-level sheet and sum CO2 by BA → ISO."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    print(f"Reading {EGRID_FILE}...")

    # The plant-level sheet is typically named PLNT23 or similar
    wb = openpyxl.load_workbook(EGRID_FILE, read_only=True, data_only=True)

    # Find the plant sheet
    plant_sheet = None
    for name in wb.sheetnames:
        if 'PLNT' in name.upper() or 'plant' in name.lower():
            plant_sheet = name
            break

    if not plant_sheet:
        print(f"Available sheets: {wb.sheetnames}")
        print("ERROR: Could not find plant-level sheet (expected PLNT23 or similar)")
        sys.exit(1)

    print(f"Using sheet: {plant_sheet}")
    ws = wb[plant_sheet]

    # Find column indices from header row
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=False):
        for cell in row:
            if cell.value:
                headers[str(cell.value).strip().upper()] = cell.column - 1

    # Key columns we need
    # BACODE = balancing authority code
    # PLCO2AN = plant annual CO2 emissions (short tons)
    # PLNGENAN = plant annual net generation (MWh)
    # PLFUELCT = plant primary fuel category
    ba_col = None
    co2_col = None
    gen_col = None
    fuel_col = None

    for key, idx in headers.items():
        if 'BACODE' in key or key == 'BACODE':
            ba_col = idx
        if 'PLCO2AN' in key or key == 'PLCO2AN':
            co2_col = idx
        if 'PLNGENAN' in key or key == 'PLNGENAN':
            gen_col = idx
        if 'PLFUELCT' in key:
            fuel_col = idx

    if ba_col is None or co2_col is None:
        # Try alternate column names
        print(f"Header keys found: {list(headers.keys())[:30]}")
        print("Looking for BACODE and PLCO2AN columns...")
        # Search more broadly
        for key, idx in headers.items():
            if 'BA' in key and 'CODE' in key:
                ba_col = idx
            if 'CO2' in key and 'AN' in key and 'PL' in key:
                co2_col = idx
            if 'NEGEN' in key or ('GEN' in key and 'AN' in key and 'PL' in key):
                gen_col = idx

    if ba_col is None or co2_col is None:
        print(f"ERROR: Could not find required columns.")
        print(f"  BA column index: {ba_col}")
        print(f"  CO2 column index: {co2_col}")
        print(f"  Available headers: {list(headers.keys())[:50]}")
        sys.exit(1)

    print(f"Column indices — BA: {ba_col}, CO2: {co2_col}, Gen: {gen_col}, Fuel: {fuel_col}")

    # Aggregate by ISO
    iso_co2_short_tons = {iso: 0.0 for iso in BA_TO_ISO.values()}
    iso_gen_mwh = {iso: 0.0 for iso in BA_TO_ISO.values()}
    iso_plant_count = {iso: 0 for iso in BA_TO_ISO.values()}
    fuel_breakdown = {iso: {} for iso in BA_TO_ISO.values()}

    total_plants = 0
    matched_plants = 0

    for row in ws.iter_rows(min_row=3, values_only=True):  # Skip header rows
        total_plants += 1
        ba_code = row[ba_col] if ba_col < len(row) else None
        if ba_code and str(ba_code).strip() in BA_TO_ISO:
            iso = BA_TO_ISO[str(ba_code).strip()]
            co2 = row[co2_col] if co2_col < len(row) else 0
            gen = row[gen_col] if gen_col is not None and gen_col < len(row) else 0
            fuel = row[fuel_col] if fuel_col is not None and fuel_col < len(row) else 'Unknown'

            if co2 and isinstance(co2, (int, float)):
                iso_co2_short_tons[iso] += co2
                matched_plants += 1
                iso_plant_count[iso] += 1

            if gen and isinstance(gen, (int, float)):
                iso_gen_mwh[iso] += gen

            if fuel:
                fuel = str(fuel).strip()
                if fuel not in fuel_breakdown[iso]:
                    fuel_breakdown[iso][fuel] = {'co2_short_tons': 0, 'gen_mwh': 0}
                if co2 and isinstance(co2, (int, float)):
                    fuel_breakdown[iso][fuel]['co2_short_tons'] += co2
                if gen and isinstance(gen, (int, float)):
                    fuel_breakdown[iso][fuel]['gen_mwh'] += gen

    wb.close()

    print(f"\nProcessed {total_plants} plants, {matched_plants} matched to 7 ISOs")

    # Convert and build output
    results = {}
    print(f"\n{'ISO':<8} {'Plants':>7} {'CO2 (M short tons)':>20} {'CO2 (M metric tons)':>22} {'Gen (TWh)':>10}")
    print("-" * 70)

    for iso in ['CAISO', 'ERCOT', 'PJM', 'NYISO', 'NEISO', 'MISO', 'SPP']:
        co2_st = iso_co2_short_tons[iso]
        co2_mt = co2_st * SHORT_TONS_TO_METRIC_TONS
        gen_twh = iso_gen_mwh[iso] / 1e6

        results[iso] = {
            'co2_short_tons': round(co2_st),
            'co2_metric_tons': round(co2_mt),
            'co2_million_metric_tons': round(co2_mt / 1e6, 2),
            'generation_twh': round(gen_twh, 1),
            'plant_count': iso_plant_count[iso],
            'source': 'EPA eGRID 2023 plant-level (PLNT23)',
            'at_trajectory': {
                '2030_cap_mt': round(co2_mt * 0.43),
                '2035_cap_mt': round(co2_mt * 0.18),
                '2040_cap_mt': round(co2_mt * 0.12),
                '2045_cap_mt': round(co2_mt * 0.06),
                '2050_cap_mt': 0,
            },
            'fuel_breakdown': {
                fuel: {
                    'co2_million_metric_tons': round(data['co2_short_tons'] * SHORT_TONS_TO_METRIC_TONS / 1e6, 2),
                    'generation_twh': round(data['gen_mwh'] / 1e6, 1),
                }
                for fuel, data in sorted(fuel_breakdown[iso].items())
            }
        }

        print(f"{iso:<8} {iso_plant_count[iso]:>7} {co2_st/1e6:>20.2f} {co2_mt/1e6:>22.2f} {gen_twh:>10.1f}")

    total_mt = sum(r['co2_million_metric_tons'] for r in results.values())
    print(f"\n{'TOTAL':<8} {'':>7} {'':>20} {total_mt:>22.2f}")

    # Save results
    with open(OUTPUT_FILE, 'w') as f:
        json.dump(results, f, indent=2)

    print(f"\nSaved to: {OUTPUT_FILE}")
    return results


if __name__ == '__main__':
    if download_egrid():
        extract_plant_emissions()
    else:
        print("\nTo run manually after downloading:")
        print(f"  python {__file__}")
